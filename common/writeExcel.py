#！ /usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2019/11/9 0009 下午 23:38 
# @Author : aiminhu
# @File : writeExcel.py 
# @Software: PyCharm
'''
功能：
    1.知道excel,读取数据
    2.复制excel对象
    3.写入实际结果
PS:文件名和类名一致时，需要注意
'''
import xlrd
import openpyxl
import os


class writeExcel():
    dir = 'testData'
    # excel_dir = os.path.dirname(os.getcwd()) + "\\" + dir  #在当前目录下执行，可正确获取
    excel_dir = os.getcwd() + "\\" + dir
    # print(excel_dir)
    wb = openpyxl.load_workbook(excel_dir + '\\' + 'data.xlsx')
    sheetnames = wb.sheetnames
    # print(sheetnames)
    # # excel_dir = os.getcwd() + "\\" + dir
    #
    # workbook = xlrd.open_workbook(excel_dir + '\\' + 'data.xlsx')
    # # 根据sheet索引或者名称获取sheet内容
    # sheet1 = workbook.sheet_by_index(0)
    # urlSheet = workbook.sheet_by_name('urlSheet')
    # paramSheet = workbook.sheet_by_name('paramSheet')
    # assertSheet = workbook.sheet_by_name('assertSheet')
    # rownum = urlSheet.nrows


    def writeResult(self,id,real,result):
        #打开一个Excel
        # wb = openpyxl.Workbook()
        # sheetnames = wb.sheetnames
        # print(sheetnames)
        # #打开名字为assertSheet的sheet
        ws = self.wb['assertSheet']
        # #写入数据，id是行和编号，real是实际返回值，result是结果
        ws.cell(id+1,3).value = real
        ws.cell(id+1,4).value = result
        #保存
        self.wb.save(self.excel_dir + '\\' + 'data.xlsx')


wr = writeExcel()
# wr.writeResult(1,'0','success')